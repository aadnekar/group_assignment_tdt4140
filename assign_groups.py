#!/bin/python3

import math
import argparse
import os.path
from openpyxl import Workbook, load_workbook

parser = argparse.ArgumentParser(
    description="Group assignment tool created for TDT4140 at NTNU, authored by Ådne Karstad"
)

parser.add_argument(
    '-n', '--number_of_groups',
    type=int,
    required=False,
    help='total number of groups that students may be assigned to.'
)
parser.add_argument(
    '-Gs', "--group_size",
    type=int,
    required=False,
    help="the maximum number of students that may be assigned to a group. Set either --number_of_groups or --group_size",
)
parser.add_argument(
    "-o", '--file_destination',
    type=str,
    default="assigned_groups",
    help="Filename of output, e.g. '-o result', will output a file result.xlsx",
)
parser.add_argument("-f", '--file_source', type=str, help="Path to a file to process")
parser.add_argument("--filter", action='store_true', default=False)

args = parser.parse_args()

ATTRIBUTES_MAP = {
    "B": "username",
    "C": "name",
    "D": "email",
    "E": "provided_email",
    "F": "programming_experience",
    "G": "completed_it1901",
    "J": "reference_group",
    "I": "available_on_compulsory_dates",
    "H": "granted_permission_for_deliveries",
    "K": "program",
}


ATTRIBUTES_IN_MAP = {
    # "group": -1,
    # "username": -1,
    "name": 4,
    "email": 3,
    "provided_email": 5,
    "programming_experience": 7,
    "completed_it1901": 6,
    "reference_group": 10,
    "available_on_compulsory_dates": 8,
    "granted_permission_for_deliveries": 9,
    "program": 11,
    "time": 2
}

ATTRIBUTES_TO_PRETTY_FORMAT = {
    "group": "Gruppe",
    "username": "Brukernavn",
    "name": "Navn",
    "email": "Email",
    "provided_email": "Gitt email",
    "programming_experience": "Programmeringsferdighet",
    "completed_it1901": "Gjennomført IT1901",
    "reference_group": "Ønsker å være i referansegruppe",
    "available_on_compulsory_dates": "Tilgjengelig under produktgjennomgang",
    "granted_permission_for_deliveries": "Gitt tillatelse for bruk av anonymiserte leveranser",
    "program": "Studieprogram",
}


class Student(object):
    """ Object to represent the required fields of students in TDT4140 """

    def __init__(self, **kwargs):
        """
            Values that should be assigned to a student object should be adjusted in the directories
            at the beginning of the file.
        """

        self.__dict__.update(kwargs)
        
        if "email" in kwargs.keys():
            self.username = self.get_username(kwargs["email"])
        

    def __eq__(self, other):
        """ Check if the student has the same student email """
        return self.email == other.email

    def is_newer_than(self, other):
        """ True if self is new than other """
        return self.time > other.time


    @staticmethod
    def get_username(email):
        return email.split("@")[0] if email else ""

    def get_programming_experience(self):
        return self.programming_experience if self.programming_experience else 0

    def __getattribute__(self, name: str):
        return super().__getattribute__(name)

def store_students_in_list(workbook, students=None):
    """ Temporal storage for students to be sorted """
    if not students:
        students = []

    for i, student in enumerate(workbook.active.values):
        # Do not include the name of columns row
        if i == 0:
            continue

        student_attr = {}
        for key, value in zip(ATTRIBUTES_IN_MAP.keys(), ATTRIBUTES_IN_MAP.values()):
            student_attr[key] = student[value]

        student = Student(**student_attr)

        if args.filter and student in students:
            """ Replace with old instance or ignore new one """
            other_stud = students[students.index(student)]

            if student.is_newer_than(other_stud):
                students[students.index(other_stud)] = student
        else:
            students.append(student)

    return students


def convert_answer_to_bool(answer):
    if answer.lower() == "ja":
        return True
    else:
        return False


def sort_students(students):
    """
    student: list of student objects

    sorted on attributes:
        - stud.taken_it1901
        - stud.prog_skill

    Results in the topological sorting with the highest priority of grant_anonym,
    and least significant of programming skill.
    """

    students = sorted(
        students, key=lambda stud: getattr(stud, "programming_experience")
    )
    students = sorted(
        students,
        key=lambda stud: convert_answer_to_bool(getattr(stud, "completed_it1901")),
        reverse=True,
    )

    return students


def check_if_duplicate(student, students):
    """ Keeps only the newest student with the same email """

    return student in students
        

def write_column_names(sheet):
    """ Add column name to row 1 """
    for key, attribute in zip(ATTRIBUTES_MAP.keys(), ATTRIBUTES_MAP.values()):
        sheet[f"{key}1"] = f"{ATTRIBUTES_TO_PRETTY_FORMAT[attribute]}"


def write_student_to_sheet(sheet, row, group_index, student=None):
    """ Write a single student to the active worksheet """
    sheet[f"A{row}"] = f"Group {group_index + 1}"
    for key, attribute in zip(ATTRIBUTES_MAP.keys(), ATTRIBUTES_MAP.values()):
        sheet[f"{key}{row}"] = f"{getattr(student, attribute)}"


def assign_students_to_groups(
    sheet, number_of_groups, students, group_size=8, offset_value=2
):
    student_index = 0  # Represent the index of student to be appeneded to the group
    group_index = 0  # Represent the group to append a student to

    while len(students) > 0:
        if group_index % number_of_groups == 0:

            group_index = 0
            student_index += 1

        write_student_to_sheet(
            sheet,
            row=calculate_row(group_index, group_size, offset_value, student_index),
            group_index=group_index,
            student=students.pop(),
        )

        group_index += 1

    return sheet

def calculate_row(group_index, group_size, offset_value=2, student_index=0):
    """
    Calculate the relevant row
    Arguments:
        - group_index -> what group is currently referenced
        - group_size -> what is the maximum size of groups
        - offset_value -> what row does group 1 start at
        - student_index -> what row in a given group is referenced
    """
    return offset_value + group_index * (group_size + 1) + student_index


def main():
    """ Main function """

    # You cannot both set the number_of_groups and group_size
    # If you do, the number_of_groups argument trumphs the other.
    if args.number_of_groups:
        number_of_groups = args.number_of_groups
        group_size = -1
    elif args.group_size:
        group_size = args.group_size
        number_of_groups = -1
    else:
        raise argparse.ArgumentError(message='You must either provide number_of_groups or group_size.', argument=None)

    if not os.path.isfile(args.file_source):
        raise FileNotFoundError(
            f"Cannot locate the file that you have provided {args.file_source}"
        )

    loaded_workbook = load_workbook(filename=f"{args.file_source}")

    students = store_students_in_list(workbook=loaded_workbook)
    students = sort_students(students)

    number_of_students = len(students)

    # You either have to calculate one or the other.
    if group_size > 0:
        number_of_groups = math.ceil(number_of_students / group_size)
    else:
        group_size = math.ceil(number_of_students / number_of_groups)

    new_workbook = Workbook()
    sheet = new_workbook.active

    write_column_names(sheet)

    assign_students_to_groups(
        sheet=sheet, number_of_groups=number_of_groups, group_size=group_size, students=students
    )

    new_workbook.save(filename=f"{args.file_destination}.xlsx")


if __name__ == "__main__":
    main()
